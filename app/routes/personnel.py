from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, jsonify
from flask_login import login_required, current_user
from app.extensions import db
from app.models.user import Role
from app.models.unit import DonVi
from app.models.personnel import QuanNhan, CapBac, HocHam, HocVi, DoiTuong
from app.models.certificate import ChungChi, LoaiChungChi
from app.models.catalog import ChucVuOption, CapBacOption, DoiTuongOption
from app.models.evaluation import DanhGiaHangNam
from app.models.transfer import ChuyenDonVi, TrangThaiChuyen
from app.utils.decorators import unit_user_required
from app.utils.file_upload import save_upload, delete_upload
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import case, collate
from sqlalchemy.orm import aliased
from sqlalchemy.exc import ProgrammingError, OperationalError, IntegrityError

personnel_bp = Blueprint('personnel', __name__)


@personnel_bp.route('/')
@login_required
@unit_user_required
def list_personnel():
    if not current_user.don_vi:
        flash('Tài khoản chưa được gán đơn vị.', 'warning')
        return redirect(url_for('dashboard.index'))

    search = request.args.get('search', '').strip()
    cap_bac_filter = request.args.get('cap_bac', '').strip()
    chuc_vu_filter = request.args.get('chuc_vu', '').strip()
    doi_tuong_filter = request.args.get('doi_tuong', '').strip()
    don_vi_truc_thuoc_filter = request.args.get('don_vi_truc_thuoc', '').strip()
    sort_by = request.args.get('sort_by', 'chuc_vu_order').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    if per_page not in (20, 50, 100):
        per_page = 20

    query = QuanNhan.query.filter_by(don_vi_id=current_user.don_vi_id, is_active=True)
    try:
        query = query.filter(QuanNhan.is_deleted == False)
    except Exception:
        pass
    if search:
        query = query.filter(QuanNhan.ho_ten.ilike(f'%{search}%'))
    if cap_bac_filter:
        query = query.filter(QuanNhan.cap_bac == cap_bac_filter)
    if chuc_vu_filter:
        query = query.filter(QuanNhan.chuc_vu == chuc_vu_filter)
    if doi_tuong_filter:
        query = query.filter(QuanNhan.doi_tuong == doi_tuong_filter)
    if don_vi_truc_thuoc_filter:
        query = query.filter(QuanNhan.don_vi_truc_thuoc == don_vi_truc_thuoc_filter)

    chuc_vu_alias = aliased(ChucVuOption)
    query = query.outerjoin(
        chuc_vu_alias,
        collate(chuc_vu_alias.ten, 'utf8mb4_unicode_ci') == collate(QuanNhan.chuc_vu, 'utf8mb4_unicode_ci')
    )

    if sort_by == 'ho_ten':
        query = query.order_by(QuanNhan.ho_ten.asc())
    elif sort_by == 'cap_bac':
        query = query.order_by(
            case((QuanNhan.cap_bac.is_(None), 1), else_=0),
            QuanNhan.cap_bac.asc(),
            QuanNhan.ho_ten.asc()
        )
    elif sort_by == 'doi_tuong':
        query = query.order_by(
            case((QuanNhan.doi_tuong.is_(None), 1), else_=0),
            QuanNhan.doi_tuong.asc(),
            QuanNhan.ho_ten.asc()
        )
    else:
        query = query.order_by(
            case((chuc_vu_alias.thu_tu.is_(None), 1), else_=0),
            chuc_vu_alias.thu_tu.asc(),
            QuanNhan.chuc_vu.asc(),
            QuanNhan.ho_ten.asc()
        )

    personnel = query.paginate(page=page, per_page=per_page, error_out=False)

    # Distinct don_vi_truc_thuoc values for this don_vi (for filter dropdown)
    don_vi_truc_thuoc_list = [
        row[0] for row in db.session.query(QuanNhan.don_vi_truc_thuoc)
        .filter(QuanNhan.don_vi_id == current_user.don_vi_id, QuanNhan.is_active == True,
                QuanNhan.don_vi_truc_thuoc.isnot(None), QuanNhan.don_vi_truc_thuoc != '')
        .distinct().order_by(QuanNhan.don_vi_truc_thuoc).all()
    ]

    return render_template('personnel/list.html',
                           personnel=personnel,
                           search=search,
                           cap_bac_filter=cap_bac_filter,
                           chuc_vu_filter=chuc_vu_filter,
                           doi_tuong_filter=doi_tuong_filter,
                           don_vi_truc_thuoc_filter=don_vi_truc_thuoc_filter,
                           sort_by=sort_by,
                           per_page=per_page,
                           cap_bac_list=_get_cap_bac_list(),
                           chuc_vu_list=[x.ten for x in _get_chuc_vu_options()],
                           doi_tuong_list=_get_doi_tuong_list(),
                           don_vi_truc_thuoc_list=don_vi_truc_thuoc_list)


def _get_chuc_vu_options():
    return ChucVuOption.query.filter_by(is_active=True).order_by(ChucVuOption.thu_tu, ChucVuOption.ten).all()


def _get_cap_bac_list():
    db_values = [x.ten for x in CapBacOption.query.filter_by(is_active=True).order_by(CapBacOption.thu_tu, CapBacOption.ten).all()]
    return db_values if db_values else [e.value for e in CapBac]


def _get_doi_tuong_list():
    db_values = [x.ten for x in DoiTuongOption.query.filter_by(is_active=True).order_by(DoiTuongOption.thu_tu, DoiTuongOption.ten).all()]
    return db_values if db_values else [e.value for e in DoiTuong]


@personnel_bp.route('/create', methods=['GET', 'POST'])
@login_required
@unit_user_required
def create_personnel():
    if not current_user.don_vi:
        flash('Tài khoản chưa được gán đơn vị.', 'warning')
        return redirect(url_for('dashboard.index'))

    if request.method == 'POST':
        ngay_sinh = None
        ns_str = request.form.get('ngay_sinh', '').strip()
        if ns_str:
            try:
                ngay_sinh = datetime.strptime(ns_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        qn = QuanNhan(
            don_vi_id=current_user.don_vi_id,
            ho_ten=request.form.get('ho_ten', '').strip(),
            cap_bac=request.form.get('cap_bac', '').strip() or None,
            chuc_danh=None,
            chuc_vu=request.form.get('chuc_vu', '').strip() or None,
            don_vi_truc_thuoc=request.form.get('don_vi_truc_thuoc', '').strip() or None,
            can_cuoc_cong_dan=request.form.get('can_cuoc_cong_dan', '').strip() or None,
            ngay_sinh=ngay_sinh,
            ngay_nhap_ngu=request.form.get('ngay_nhap_ngu', '').strip() or None,
            doi_tuong=request.form.get('doi_tuong', '').strip() or None,
            hoc_ham=request.form.get('hoc_ham', 'Không').strip(),
            hoc_vi=request.form.get('hoc_vi', 'Không').strip(),
            trinh_do_hoc_van=request.form.get('trinh_do_hoc_van', '').strip() or None,
            ngoai_ngu=request.form.get('ngoai_ngu', '').strip() or None,
            la_chi_huy='la_chi_huy' in request.form,
            la_bi_thu='la_bi_thu' in request.form,
            la_dang_vien='la_dang_vien' in request.form,
            la_doan_vien='la_doan_vien' in request.form,
            la_hoi_vien_phu_nu='la_hoi_vien_phu_nu' in request.form,
            lop=request.form.get('lop', '').strip() or None,
        )

        if not qn.ho_ten:
            flash('Họ tên không được để trống.', 'danger')
            return render_template('personnel/create.html',
                                   cap_bac_list=_get_cap_bac_list(),
                                   hoc_ham_list=[e.value for e in HocHam],
                                   hoc_vi_list=[e.value for e in HocVi],
                                   doi_tuong_list=_get_doi_tuong_list(),
                                   chuc_vu_options=_get_chuc_vu_options())

        # Kiểm tra CCCD trùng
        cccd_val = qn.can_cuoc_cong_dan
        if cccd_val:
            existing = QuanNhan.query.filter(
                QuanNhan.can_cuoc_cong_dan == cccd_val
            ).first()
            if existing:
                flash(f'Số CCCD {cccd_val} đã tồn tại (thuộc quân nhân: {existing.ho_ten} - {existing.don_vi.ten_don_vi if existing.don_vi else ""}). Vui lòng kiểm tra lại.', 'danger')
                return render_template('personnel/create.html',
                                       cap_bac_list=_get_cap_bac_list(),
                                       hoc_ham_list=[e.value for e in HocHam],
                                       hoc_vi_list=[e.value for e in HocVi],
                                       doi_tuong_list=_get_doi_tuong_list(),
                                       chuc_vu_options=_get_chuc_vu_options())

        db.session.add(qn)
        db.session.commit()
        flash(f'Đã thêm quân nhân: {qn.ho_ten}', 'success')
        return redirect(url_for('personnel.detail_personnel', id=qn.id))

    return render_template('personnel/create.html',
                           cap_bac_list=_get_cap_bac_list(),
                           hoc_ham_list=[e.value for e in HocHam],
                           hoc_vi_list=[e.value for e in HocVi],
                           doi_tuong_list=_get_doi_tuong_list(),
                           chuc_vu_options=_get_chuc_vu_options())


@personnel_bp.route('/<int:id>')
@login_required
def detail_personnel(id):
    from app.models.user import Role as _Role
    if current_user.role not in (_Role.UNIT_USER, _Role.ADMIN):
        from flask import abort
        abort(403)
    qn = QuanNhan.query.get_or_404(id)
    # Admin can view any personnel; unit users can only view their own unit
    if current_user.role != _Role.ADMIN and qn.don_vi_id != current_user.don_vi_id:
        flash('Không có quyền truy cập.', 'danger')
        return redirect(url_for('personnel.list_personnel'))

    return render_template('personnel/detail.html', qn=qn,
                           loai_chung_chi_list=[e.value for e in LoaiChungChi],
                           back_url=request.referrer or url_for('personnel.list_personnel'))


@personnel_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@unit_user_required
def edit_personnel(id):
    qn = QuanNhan.query.get_or_404(id)
    if qn.don_vi_id != current_user.don_vi_id:
        flash('Không có quyền truy cập.', 'danger')
        return redirect(url_for('personnel.list_personnel'))

    if request.method == 'POST':
        qn.ho_ten = request.form.get('ho_ten', '').strip()
        qn.cap_bac = request.form.get('cap_bac', '').strip() or None
        qn.chuc_danh = None
        qn.chuc_vu = request.form.get('chuc_vu', '').strip() or None
        qn.don_vi_truc_thuoc = request.form.get('don_vi_truc_thuoc', '').strip() or None
        qn.can_cuoc_cong_dan = request.form.get('can_cuoc_cong_dan', '').strip() or None
        qn.doi_tuong = request.form.get('doi_tuong', '').strip() or None
        qn.hoc_ham = request.form.get('hoc_ham', 'Không').strip()
        qn.hoc_vi = request.form.get('hoc_vi', 'Không').strip()
        qn.trinh_do_hoc_van = request.form.get('trinh_do_hoc_van', '').strip() or None
        qn.ngoai_ngu = request.form.get('ngoai_ngu', '').strip() or None
        qn.ngay_nhap_ngu = request.form.get('ngay_nhap_ngu', '').strip() or None
        qn.la_chi_huy = 'la_chi_huy' in request.form
        qn.la_bi_thu = 'la_bi_thu' in request.form
        qn.la_dang_vien = 'la_dang_vien' in request.form
        qn.la_doan_vien = 'la_doan_vien' in request.form
        qn.la_hoi_vien_phu_nu = 'la_hoi_vien_phu_nu' in request.form
        qn.lop = request.form.get('lop', '').strip() or None

        ns_str = request.form.get('ngay_sinh', '').strip()
        if ns_str:
            try:
                qn.ngay_sinh = datetime.strptime(ns_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        else:
            qn.ngay_sinh = None

        if not qn.ho_ten:
            flash('Họ tên không được để trống.', 'danger')
        else:
            db.session.commit()
            flash('Đã cập nhật thông tin.', 'success')
            return redirect(url_for('personnel.detail_personnel', id=qn.id))

    return render_template('personnel/edit.html', qn=qn,
                           cap_bac_list=_get_cap_bac_list(),
                           hoc_ham_list=[e.value for e in HocHam],
                           hoc_vi_list=[e.value for e in HocVi],
                           doi_tuong_list=_get_doi_tuong_list(),
                           chuc_vu_options=_get_chuc_vu_options())


@personnel_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
@unit_user_required
def delete_personnel(id):
    qn = QuanNhan.query.get_or_404(id)
    if qn.don_vi_id != current_user.don_vi_id:
        flash('Không có quyền truy cập.', 'danger')
        return redirect(url_for('personnel.list_personnel'))

    qn.is_active = False
    try:
        qn.is_deleted = True
        qn.deleted_at = datetime.utcnow()
        qn.deleted_by_id = current_user.id
    except Exception:
        pass
    db.session.commit()
    flash(f'Đã đưa vào danh sách xóa: {qn.ho_ten}', 'warning')
    return redirect(url_for('personnel.list_personnel'))


@personnel_bp.route('/<int:id>/certificate', methods=['POST'])
@login_required
@unit_user_required
def add_certificate(id):
    qn = QuanNhan.query.get_or_404(id)
    if qn.don_vi_id != current_user.don_vi_id:
        flash('Không có quyền truy cập.', 'danger')
        return redirect(url_for('personnel.list_personnel'))

    ten = request.form.get('ten_chung_chi', '').strip()
    if not ten:
        flash('Tên chứng chỉ không được để trống.', 'danger')
        return redirect(url_for('personnel.detail_personnel', id=qn.id))

    duong_dan_anh = None
    file = request.files.get('anh_minh_chung')
    if file and file.filename:
        duong_dan_anh = save_upload(file, 'certificates')

    ngay_cap = None
    nc_str = request.form.get('ngay_cap', '').strip()
    if nc_str:
        try:
            ngay_cap = datetime.strptime(nc_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    cc = ChungChi(
        quan_nhan_id=qn.id,
        loai=request.form.get('loai', LoaiChungChi.THANH_TICH_KHAC.value),
        ten_chung_chi=ten,
        so_hieu=request.form.get('so_hieu', '').strip() or None,
        ngay_cap=ngay_cap,
        co_quan_cap=request.form.get('co_quan_cap', '').strip() or None,
        duong_dan_anh=duong_dan_anh,
        ghi_chu=request.form.get('ghi_chu', '').strip() or None,
    )
    db.session.add(cc)
    db.session.commit()
    flash(f'Đã thêm: {ten}', 'success')
    return redirect(url_for('personnel.detail_personnel', id=qn.id))


@personnel_bp.route('/certificate/<int:id>/delete', methods=['POST'])
@login_required
@unit_user_required
def delete_certificate(id):
    cc = ChungChi.query.get_or_404(id)
    qn = cc.quan_nhan
    if qn.don_vi_id != current_user.don_vi_id:
        flash('Không có quyền truy cập.', 'danger')
        return redirect(url_for('personnel.list_personnel'))

    if cc.duong_dan_anh:
        delete_upload(cc.duong_dan_anh)

    db.session.delete(cc)
    db.session.commit()
    flash('Đã xóa chứng chỉ.', 'success')
    return redirect(url_for('personnel.detail_personnel', id=qn.id))


def _parse_bool(value):
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value == 1
    return str(value).strip().lower() in {'1', 'true', 'yes', 'y', 'x', 'có', 'co'}


def _parse_date(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'date'):
        try:
            return value.date()
        except Exception:
            pass
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_ngay_nhap_ngu(value):
    """Normalize ngay_nhap_ngu to MM/YYYY string, stripping time if present."""
    if value is None or value == '':
        return None
    # datetime or date object from Excel
    if isinstance(value, datetime):
        return value.strftime('%m/%Y')
    if hasattr(value, 'strftime'):
        try:
            return value.strftime('%m/%Y')
        except Exception:
            pass
    text = str(value).strip()
    # Already in correct format MM/YYYY
    import re
    if re.match(r'^\d{2}/\d{4}$', text):
        return text
    # datetime string like "2015-01-01 00:00:00" or "2015-01-01"
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(text, fmt).strftime('%m/%Y')
        except ValueError:
            continue
    # fallback: strip time part if space present
    if ' ' in text:
        text = text.split(' ')[0]
    return text or None


def _normalize_cccd(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith('.0') and text.replace('.', '', 1).isdigit():
        text = text[:-2]
    return text


@personnel_bp.route('/template')
@login_required
@unit_user_required
def download_personnel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'QuanNhan'

    lookup = wb.create_sheet('DanhMuc')
    lookup.sheet_state = 'hidden'

    cap_bac_values = _get_cap_bac_list()
    doi_tuong_values = _get_doi_tuong_list()
    hoc_ham_values = [e.value for e in HocHam]
    hoc_vi_values = [e.value for e in HocVi]
    chuc_vu_values = [x.ten for x in _get_chuc_vu_options()]
    bool_values = ['Không', 'Có']

    data_sets = [
        ('A', cap_bac_values),
        ('B', doi_tuong_values),
        ('C', hoc_ham_values),
        ('D', hoc_vi_values),
        ('E', chuc_vu_values),
        ('F', bool_values),
    ]
    for col, values in data_sets:
        for idx, value in enumerate(values, start=1):
            lookup[f'{col}{idx}'] = value

    headers = [
        'Họ và tên', 'Cấp bậc', 'Đối tượng', 'Lớp', 'Chức vụ', 'Căn cước công dân',
        'Đơn vị trực thuộc',
        'Ngày sinh', 'Ngày nhập ngũ', 'Học hàm', 'Học vị', 'Trình độ học vấn',
        'Ngoại ngữ', 'Cấp trưởng', 'Bí thư',
        'Đảng viên', 'Đoàn viên', 'Hội viên phụ nữ',
    ]
    ws.append(headers)
    ws.append([
        'Nguyễn Văn A', 'Trung úy', 'Giảng viên', '', 'Trợ lý', '012345678901',
        'Đại đội 1', '15/01/1990', '09/2015', 'Không', 'Thạc sĩ', '12/12',
        'Anh B2', 'Không', 'Không', 'Không', 'Không', 'Không',
    ])

    max_row = 500
    validations = [
        ('B', f"=DanhMuc!$A$1:$A${max(1, len(cap_bac_values))}"),
        ('C', f"=DanhMuc!$B$1:$B${max(1, len(doi_tuong_values))}"),
        ('J', f"=DanhMuc!$C$1:$C${max(1, len(hoc_ham_values))}"),
        ('K', f"=DanhMuc!$D$1:$D${max(1, len(hoc_vi_values))}"),
        ('E', f"=DanhMuc!$E$1:$E${max(1, len(chuc_vu_values))}"),
        ('N', "=DanhMuc!$F$1:$F$2"),
        ('O', "=DanhMuc!$F$1:$F$2"),
        ('P', "=DanhMuc!$F$1:$F$2"),
        ('Q', "=DanhMuc!$F$1:$F$2"),
        ('R', "=DanhMuc!$F$1:$F$2"),
    ]
    for col, formula in validations:
        dv = DataValidation(type='list', formula1=formula, allow_blank=True)
        dv.prompt = 'Chọn giá trị từ danh sách'
        ws.add_data_validation(dv)
        dv.add(f'{col}2:{col}{max_row}')

    for _ws in wb.worksheets:
        _ws.page_setup.paperSize = 9
        _ws.page_setup.orientation = 'landscape'
        _ws.page_setup.fitToPage = True
        _ws.page_setup.fitToWidth = 1
        _ws.page_setup.fitToHeight = 0
        _ws.sheet_properties.pageSetUpPr.fitToPage = True

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='mau_quan_nhan.xlsx',
    )


@personnel_bp.route('/import', methods=['POST'])
@login_required
@unit_user_required
def import_personnel_excel():
    if not current_user.don_vi:
        flash('Tài khoản chưa được gán đơn vị.', 'warning')
        return redirect(url_for('personnel.list_personnel'))

    file = request.files.get('excel_file')
    if not file or not file.filename:
        flash('Vui lòng chọn file Excel (.xlsx).', 'danger')
        return redirect(url_for('personnel.list_personnel'))

    if not file.filename.lower().endswith('.xlsx'):
        flash('Chỉ hỗ trợ file .xlsx.', 'danger')
        return redirect(url_for('personnel.list_personnel'))

    try:
        wb = load_workbook(file, data_only=True)
    except Exception:
        flash('Không thể đọc file Excel. Vui lòng kiểm tra định dạng.', 'danger')
        return redirect(url_for('personnel.list_personnel'))

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        flash('File Excel trống.', 'warning')
        return redirect(url_for('personnel.list_personnel'))

    header_row = [str(c).strip() if c is not None else '' for c in rows[0]]

    # Map from Vietnamese display names (in template) to internal field names
    VIET_TO_FIELD = {
        'họ và tên': 'ho_ten',
        'ho va ten': 'ho_ten',
        'họ tên': 'ho_ten',
        'ho ten': 'ho_ten',
        'ho_ten': 'ho_ten',
        'cấp bậc': 'cap_bac',
        'cap bac': 'cap_bac',
        'cap_bac': 'cap_bac',
        'đối tượng': 'doi_tuong',
        'doi tuong': 'doi_tuong',
        'doi_tuong': 'doi_tuong',
        'lớp': 'lop',
        'lop': 'lop',
        'chức vụ': 'chuc_vu',
        'chuc vu': 'chuc_vu',
        'chuc_vu': 'chuc_vu',
        'căn cước công dân': 'can_cuoc_cong_dan',
        'can cuoc cong dan': 'can_cuoc_cong_dan',
        'can_cuoc_cong_dan': 'can_cuoc_cong_dan',
        'cccd': 'can_cuoc_cong_dan',
        'đơn vị trực thuộc': 'don_vi_truc_thuoc',
        'don vi truc thuoc': 'don_vi_truc_thuoc',
        'don_vi_truc_thuoc': 'don_vi_truc_thuoc',
        'ngày sinh': 'ngay_sinh',
        'ngay sinh': 'ngay_sinh',
        'ngay_sinh': 'ngay_sinh',
        'ngày nhập ngũ': 'ngay_nhap_ngu',
        'nhập ngũ': 'ngay_nhap_ngu',
        'ngay nhap ngu': 'ngay_nhap_ngu',
        'ngay_nhap_ngu': 'ngay_nhap_ngu',
        'học hàm': 'hoc_ham',
        'hoc ham': 'hoc_ham',
        'hoc_ham': 'hoc_ham',
        'học vị': 'hoc_vi',
        'hoc vi': 'hoc_vi',
        'hoc_vi': 'hoc_vi',
        'trình độ học vấn': 'trinh_do_hoc_van',
        'trinh do hoc van': 'trinh_do_hoc_van',
        'trinh_do_hoc_van': 'trinh_do_hoc_van',
        'trình độ': 'trinh_do_hoc_van',
        'ngoại ngữ': 'ngoai_ngu',
        'ngoai ngu': 'ngoai_ngu',
        'ngoai_ngu': 'ngoai_ngu',
        'cấp trưởng': 'la_chi_huy',
        'cap truong': 'la_chi_huy',
        'la_chi_huy': 'la_chi_huy',
        'bí thư': 'la_bi_thu',
        'bi thu': 'la_bi_thu',
        'la_bi_thu': 'la_bi_thu',
        'đảng viên': 'la_dang_vien',
        'dang vien': 'la_dang_vien',
        'la_dang_vien': 'la_dang_vien',
        'đoàn viên': 'la_doan_vien',
        'doan vien': 'la_doan_vien',
        'la_doan_vien': 'la_doan_vien',
        'hội viên phụ nữ': 'la_hoi_vien_phu_nu',
        'hoi vien phu nu': 'la_hoi_vien_phu_nu',
        'la_hoi_vien_phu_nu': 'la_hoi_vien_phu_nu',
        'phụ nữ': 'la_hoi_vien_phu_nu',
    }

    # Normalize headers: strip, lowercase, map to field names
    headers = {}
    for idx, h in enumerate(header_row):
        if not h:
            continue
        import unicodedata
        h_lower = h.lower().strip()
        field = VIET_TO_FIELD.get(h_lower)
        if not field:
            # Try removing diacritics
            h_ascii = ''.join(
                c for c in unicodedata.normalize('NFD', h_lower)
                if unicodedata.category(c) != 'Mn'
            ).strip()
            field = VIET_TO_FIELD.get(h_ascii, h_lower.replace(' ', '_'))
        headers[field] = idx

    required_cols = ['ho_ten']
    optional_cols = [
        'cap_bac', 'doi_tuong', 'lop', 'chuc_vu', 'can_cuoc_cong_dan',
        'don_vi_truc_thuoc',
        'ngay_sinh', 'ngay_nhap_ngu', 'hoc_ham', 'hoc_vi', 'trinh_do_hoc_van',
        'ngoai_ngu', 'la_chi_huy', 'la_bi_thu',
        'la_dang_vien', 'la_doan_vien', 'la_hoi_vien_phu_nu',
    ]
    missing = [c for c in required_cols if c not in headers]
    if missing:
        found = list(headers.keys())[:8]
        flash(
            f'Thiếu cột bắt buộc: {", ".join(missing)}. '
            f'Các cột nhận diện được: {", ".join(found)}. '
            f'Vui lòng dùng file mẫu hoặc đặt đúng tên cột.',
            'danger'
        )
        return redirect(url_for('personnel.list_personnel'))

    created = 0
    skipped = 0
    errors = 0
    replaced = 0
    error_details = []

    cap_bac_values = set(_get_cap_bac_list())
    doi_tuong_values = {e.value for e in DoiTuong}
    hoc_ham_values = {e.value for e in HocHam}
    hoc_vi_values = {e.value for e in HocVi}

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            ho_ten = row[headers['ho_ten']] if headers.get('ho_ten') is not None else None
            ho_ten = str(ho_ten).strip() if ho_ten is not None else ''
            if not ho_ten or ho_ten.lower() in ('none', 'nan', ''):
                skipped += 1
                continue

            cccd = None
            if 'can_cuoc_cong_dan' in headers:
                cccd = _normalize_cccd(row[headers['can_cuoc_cong_dan']])

            if cccd:
                duplicates = QuanNhan.query.filter_by(
                    don_vi_id=current_user.don_vi_id,
                    can_cuoc_cong_dan=cccd,
                    is_active=True
                ).all()
                if duplicates:
                    for duplicated_qn in duplicates:
                        duplicated_qn.is_active = False
                    replaced += len(duplicates)

            cap_bac_val = str(row[headers['cap_bac']]).strip() if headers.get('cap_bac') is not None and row[headers['cap_bac']] is not None else None
            doi_tuong_val = str(row[headers['doi_tuong']]).strip() if headers.get('doi_tuong') is not None and row[headers['doi_tuong']] is not None else None
            hoc_ham_val = str(row[headers['hoc_ham']]).strip() if headers.get('hoc_ham') is not None and row[headers['hoc_ham']] is not None else 'Không'
            hoc_vi_val = str(row[headers['hoc_vi']]).strip() if headers.get('hoc_vi') is not None and row[headers['hoc_vi']] is not None else 'Không'

            if cap_bac_val and cap_bac_val not in cap_bac_values:
                error_details.append(f'Dòng {row_idx} ({ho_ten}): Cấp bậc "{cap_bac_val}" không hợp lệ.')
                errors += 1
                continue
            if doi_tuong_val and doi_tuong_val not in doi_tuong_values:
                error_details.append(f'Dòng {row_idx} ({ho_ten}): Đối tượng "{doi_tuong_val}" không hợp lệ.')
                errors += 1
                continue
            if hoc_ham_val and hoc_ham_val not in hoc_ham_values:
                # Fallback to 'Không' instead of skipping
                hoc_ham_val = 'Không'
            if hoc_vi_val and hoc_vi_val not in hoc_vi_values:
                # Fallback to 'Không' instead of skipping
                hoc_vi_val = 'Không'

            qn = QuanNhan(
                don_vi_id=current_user.don_vi_id,
                ho_ten=ho_ten,
                cap_bac=cap_bac_val,
                doi_tuong=doi_tuong_val,
                chuc_danh=None,
                lop=str(row[headers['lop']]).strip() if headers.get('lop') is not None and row[headers['lop']] is not None else None,
                chuc_vu=str(row[headers['chuc_vu']]).strip() if headers.get('chuc_vu') is not None and row[headers['chuc_vu']] is not None else None,
                can_cuoc_cong_dan=cccd or None,
                don_vi_truc_thuoc=str(row[headers['don_vi_truc_thuoc']]).strip() if headers.get('don_vi_truc_thuoc') is not None and row[headers['don_vi_truc_thuoc']] is not None else None,
                ngay_sinh=_parse_date(row[headers['ngay_sinh']]) if headers.get('ngay_sinh') is not None else None,
                ngay_nhap_ngu=_parse_ngay_nhap_ngu(row[headers['ngay_nhap_ngu']]) if headers.get('ngay_nhap_ngu') is not None and row[headers['ngay_nhap_ngu']] is not None else None,
                hoc_ham=hoc_ham_val,
                hoc_vi=hoc_vi_val,
                trinh_do_hoc_van=str(row[headers['trinh_do_hoc_van']]).strip() if headers.get('trinh_do_hoc_van') is not None and row[headers['trinh_do_hoc_van']] is not None else None,
                ngoai_ngu=str(row[headers['ngoai_ngu']]).strip() if headers.get('ngoai_ngu') is not None and row[headers['ngoai_ngu']] is not None else None,
                la_chi_huy=_parse_bool(row[headers['la_chi_huy']]) if headers.get('la_chi_huy') is not None else False,
                la_bi_thu=_parse_bool(row[headers['la_bi_thu']]) if headers.get('la_bi_thu') is not None else False,
                la_dang_vien=_parse_bool(row[headers['la_dang_vien']]) if headers.get('la_dang_vien') is not None else False,
                la_doan_vien=_parse_bool(row[headers['la_doan_vien']]) if headers.get('la_doan_vien') is not None else False,
                la_hoi_vien_phu_nu=_parse_bool(row[headers['la_hoi_vien_phu_nu']]) if headers.get('la_hoi_vien_phu_nu') is not None else False,
            )
            db.session.add(qn)
            created += 1
        except Exception as e:
            error_details.append(f'Dòng {row_idx}: Lỗi kỹ thuật ({str(e)[:80]}).')
            errors += 1

    if created > 0:
        db.session.commit()

    msg = f'Đã nhập {created} quân nhân. Bỏ qua (trống): {skipped}. Thay thế theo CCCD: {replaced}. Lỗi: {errors}.'
    flash(msg, 'success' if created > 0 else 'warning')
    if error_details:
        flash('Chi tiết lỗi: ' + ' | '.join(error_details[:10]) + (f' ... và {len(error_details)-10} lỗi khác.' if len(error_details) > 10 else ''), 'danger')
    return redirect(url_for('personnel.list_personnel'))


@personnel_bp.route('/evaluations')
@login_required
@unit_user_required
def annual_evaluations():
    if not current_user.don_vi:
        flash('Tài khoản chưa được gán đơn vị.', 'warning')
        return redirect(url_for('dashboard.index'))

    nam_hoc = request.args.get('nam_hoc', '').strip()
    if not nam_hoc:
        current_year = datetime.now().year
        nam_hoc = f'{current_year}-{current_year + 1}'

    personnel = QuanNhan.query.filter_by(
        don_vi_id=current_user.don_vi_id,
        is_active=True
    ).order_by(QuanNhan.ho_ten.asc()).all()

    qn_ids = [qn.id for qn in personnel]
    existing_rows = {}
    nam_hoc_list = []
    migration_missing = False
    try:
        if qn_ids:
            rows = DanhGiaHangNam.query.filter(
                DanhGiaHangNam.nam_hoc == nam_hoc,
                DanhGiaHangNam.quan_nhan_id.in_(qn_ids)
            ).all()
            existing_rows = {row.quan_nhan_id: row for row in rows}

        nam_hoc_list = [row[0] for row in db.session.query(DanhGiaHangNam.nam_hoc).filter(
            DanhGiaHangNam.don_vi_id == current_user.don_vi_id
        ).distinct().order_by(DanhGiaHangNam.nam_hoc.desc()).all()]
    except (ProgrammingError, OperationalError):
        db.session.rollback()
        migration_missing = True
        flash('Thiếu bảng đánh giá hằng năm. Vui lòng chạy: flask db upgrade', 'danger')

    if nam_hoc not in nam_hoc_list:
        nam_hoc_list = [nam_hoc] + nam_hoc_list

    return render_template('personnel/evaluation_list.html',
                           personnel=personnel,
                           existing_rows=existing_rows,
                           nam_hoc=nam_hoc,
                           nam_hoc_list=nam_hoc_list,
                           migration_missing=migration_missing,
                           dang_vien_choices=DanhGiaHangNam.XEP_LOAI_DANG_VIEN_CHOICES,
                           can_bo_choices=DanhGiaHangNam.XEP_LOAI_CAN_BO_CHOICES,
                           doan_vien_choices=DanhGiaHangNam.XEP_LOAI_DOAN_VIEN_CHOICES,
                           phu_nu_choices=DanhGiaHangNam.XEP_LOAI_PHU_NU_CHOICES)


@personnel_bp.route('/evaluations/save', methods=['POST'])
@login_required
@unit_user_required
def save_annual_evaluations():
    if not current_user.don_vi:
        flash('Tài khoản chưa được gán đơn vị.', 'warning')
        return redirect(url_for('dashboard.index'))

    nam_hoc = request.form.get('nam_hoc', '').strip()
    if not nam_hoc:
        flash('Năm học không được để trống.', 'danger')
        return redirect(url_for('personnel.annual_evaluations'))

    personnel = QuanNhan.query.filter_by(don_vi_id=current_user.don_vi_id, is_active=True).all()
    if not personnel:
        flash('Đơn vị chưa có quân nhân để đánh giá.', 'warning')
        return redirect(url_for('personnel.annual_evaluations', nam_hoc=nam_hoc))

    qn_ids = [qn.id for qn in personnel]
    try:
        existing_rows = DanhGiaHangNam.query.filter(
            DanhGiaHangNam.nam_hoc == nam_hoc,
            DanhGiaHangNam.quan_nhan_id.in_(qn_ids)
        ).all()
    except (ProgrammingError, OperationalError):
        db.session.rollback()
        flash('Thiếu bảng đánh giá hằng năm. Vui lòng chạy: flask db upgrade', 'danger')
        return redirect(url_for('personnel.annual_evaluations', nam_hoc=nam_hoc))
    existing_map = {row.quan_nhan_id: row for row in existing_rows}

    apply_all_dang_vien = request.form.get('bulk_xep_loai_dang_vien', '').strip()
    apply_all_can_bo = request.form.get('bulk_xep_loai_can_bo', '').strip()

    if apply_all_dang_vien and apply_all_dang_vien not in DanhGiaHangNam.XEP_LOAI_DANG_VIEN_CHOICES:
        flash('Giá trị áp dụng hàng loạt (đảng viên) không hợp lệ.', 'danger')
        return redirect(url_for('personnel.annual_evaluations', nam_hoc=nam_hoc))

    if apply_all_can_bo and apply_all_can_bo not in DanhGiaHangNam.XEP_LOAI_CAN_BO_CHOICES:
        flash('Giá trị áp dụng hàng loạt (cán bộ) không hợp lệ.', 'danger')
        return redirect(url_for('personnel.annual_evaluations', nam_hoc=nam_hoc))

    values_map = {}
    missing_names = []
    for qn in personnel:
        dang_vien_val = request.form.get(f'xep_loai_dang_vien_{qn.id}', '').strip() or apply_all_dang_vien
        can_bo_val = request.form.get(f'xep_loai_can_bo_{qn.id}', '').strip() or apply_all_can_bo
        doan_vien_val = request.form.get(f'xep_loai_doan_vien_{qn.id}', '').strip() or None
        phu_nu_val = request.form.get(f'xep_loai_phu_nu_{qn.id}', '').strip() or None

        if not dang_vien_val or not can_bo_val:
            missing_names.append(qn.ho_ten)
            continue

        if dang_vien_val not in DanhGiaHangNam.XEP_LOAI_DANG_VIEN_CHOICES:
            flash(f'Giá trị xếp loại đảng viên không hợp lệ cho {qn.ho_ten}.', 'danger')
            return redirect(url_for('personnel.annual_evaluations', nam_hoc=nam_hoc))
        if can_bo_val not in DanhGiaHangNam.XEP_LOAI_CAN_BO_CHOICES:
            flash(f'Giá trị xếp loại cán bộ không hợp lệ cho {qn.ho_ten}.', 'danger')
            return redirect(url_for('personnel.annual_evaluations', nam_hoc=nam_hoc))

        values_map[qn.id] = (dang_vien_val, can_bo_val, doan_vien_val, phu_nu_val)

    if missing_names:
        flash(f'Chưa đánh giá đủ toàn bộ quân nhân. Thiếu: {len(missing_names)} người.', 'danger')
        return redirect(url_for('personnel.annual_evaluations', nam_hoc=nam_hoc))

    saved = 0
    for qn in personnel:
        dang_vien_val, can_bo_val, doan_vien_val, phu_nu_val = values_map[qn.id]

        row = existing_map.get(qn.id)
        if not row:
            row = DanhGiaHangNam.query.filter_by(
                quan_nhan_id=qn.id,
                nam_hoc=nam_hoc,
            ).first()
        if not row:
            row = DanhGiaHangNam(
                quan_nhan_id=qn.id,
                don_vi_id=current_user.don_vi_id,
                nam_hoc=nam_hoc,
            )
            db.session.add(row)

        row.xep_loai_dang_vien = dang_vien_val
        row.xep_loai_can_bo = can_bo_val
        row.xep_loai_doan_vien = doan_vien_val
        row.xep_loai_phu_nu = phu_nu_val
        row.nguoi_cap_nhat_id = current_user.id
        saved += 1

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()

        # Retry safely in case another request inserted same (quan_nhan_id, nam_hoc)
        saved = 0
        for qn in personnel:
            dang_vien_val, can_bo_val, doan_vien_val, phu_nu_val = values_map[qn.id]
            row = DanhGiaHangNam.query.filter_by(
                quan_nhan_id=qn.id,
                nam_hoc=nam_hoc,
            ).first()
            if not row:
                row = DanhGiaHangNam(
                    quan_nhan_id=qn.id,
                    don_vi_id=current_user.don_vi_id,
                    nam_hoc=nam_hoc,
                )
                db.session.add(row)

            row.xep_loai_dang_vien = dang_vien_val
            row.xep_loai_can_bo = can_bo_val
            row.xep_loai_doan_vien = doan_vien_val
            row.xep_loai_phu_nu = phu_nu_val
            row.nguoi_cap_nhat_id = current_user.id
            saved += 1

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash('Có xung đột dữ liệu khi lưu đánh giá. Vui lòng tải lại trang và lưu lại.', 'danger')
            return redirect(url_for('personnel.annual_evaluations', nam_hoc=nam_hoc))

    flash(f'Đã lưu {saved} đánh giá năm học {nam_hoc}.', 'success')
    return redirect(url_for('personnel.annual_evaluations', nam_hoc=nam_hoc))


# ─────────────────────────────────────────────────────────────────────────────
#  CHUYỂN ĐƠN VỊ
# ─────────────────────────────────────────────────────────────────────────────

@personnel_bp.route('/<int:id>/transfer', methods=['GET', 'POST'])
@login_required
@unit_user_required
def transfer_personnel(id):
    """Source unit initiates a transfer request."""
    qn = QuanNhan.query.get_or_404(id)
    if qn.don_vi_id != current_user.don_vi_id:
        flash('Bạn không có quyền thao tác với quân nhân này.', 'danger')
        return redirect(url_for('personnel.list_personnel'))

    # Block if there is already a pending transfer for this person
    existing = ChuyenDonVi.query.filter_by(
        quan_nhan_id=id, trang_thai=TrangThaiChuyen.PENDING
    ).first()
    if existing:
        flash('Quân nhân này đang có yêu cầu chuyển đơn vị chờ xác nhận.', 'warning')
        return redirect(url_for('personnel.detail_personnel', id=id))

    all_units = DonVi.query.filter(
        DonVi.is_active == True,
        DonVi.id != current_user.don_vi_id
    ).order_by(DonVi.ten_don_vi).all()

    if request.method == 'POST':
        don_vi_dich_id = request.form.get('don_vi_dich_id', type=int)
        ly_do = request.form.get('ly_do', '').strip()

        if not don_vi_dich_id:
            flash('Vui lòng chọn đơn vị tiếp nhận.', 'warning')
            return render_template('personnel/transfer.html', qn=qn, all_units=all_units)

        don_vi_dich = DonVi.query.get(don_vi_dich_id)
        if not don_vi_dich:
            flash('Đơn vị tiếp nhận không tồn tại.', 'danger')
            return render_template('personnel/transfer.html', qn=qn, all_units=all_units)

        chuyen = ChuyenDonVi(
            quan_nhan_id=id,
            don_vi_nguon_id=current_user.don_vi_id,
            don_vi_dich_id=don_vi_dich_id,
            nguoi_tao_id=current_user.id,
            trang_thai=TrangThaiChuyen.PENDING,
            ly_do=ly_do,
        )
        db.session.add(chuyen)
        db.session.commit()
        flash(f'Đã gửi yêu cầu chuyển {qn.ho_ten} sang {don_vi_dich.ten_don_vi}. Chờ đơn vị tiếp nhận xác nhận.', 'success')
        return redirect(url_for('personnel.detail_personnel', id=id))

    return render_template('personnel/transfer.html', qn=qn, all_units=all_units)


@personnel_bp.route('/transfers/incoming')
@login_required
@unit_user_required
def incoming_transfers():
    """Target unit views pending incoming transfer requests."""
    pending = ChuyenDonVi.query.filter_by(
        don_vi_dich_id=current_user.don_vi_id,
        trang_thai=TrangThaiChuyen.PENDING,
    ).order_by(ChuyenDonVi.ngay_tao.desc()).all()

    history = ChuyenDonVi.query.filter(
        ChuyenDonVi.don_vi_dich_id == current_user.don_vi_id,
        ChuyenDonVi.trang_thai != TrangThaiChuyen.PENDING,
    ).order_by(ChuyenDonVi.ngay_xu_ly.desc()).limit(20).all()

    outgoing = ChuyenDonVi.query.filter_by(
        don_vi_nguon_id=current_user.don_vi_id,
    ).order_by(ChuyenDonVi.ngay_tao.desc()).limit(20).all()

    return render_template(
        'personnel/transfers.html',
        pending=pending,
        history=history,
        outgoing=outgoing,
        TrangThaiChuyen=TrangThaiChuyen,
    )


@personnel_bp.route('/transfers/<int:transfer_id>/confirm', methods=['POST'])
@login_required
@unit_user_required
def confirm_transfer(transfer_id):
    """Target unit confirms the transfer — moves QuanNhan to their unit."""
    chuyen = ChuyenDonVi.query.get_or_404(transfer_id)

    if chuyen.don_vi_dich_id != current_user.don_vi_id:
        flash('Bạn không có quyền xác nhận yêu cầu này.', 'danger')
        return redirect(url_for('personnel.incoming_transfers'))

    if chuyen.trang_thai != TrangThaiChuyen.PENDING:
        flash('Yêu cầu này đã được xử lý rồi.', 'warning')
        return redirect(url_for('personnel.incoming_transfers'))

    ghi_chu = request.form.get('ghi_chu', '').strip()

    # Explicitly reload QuanNhan to ensure it is tracked by the current session
    qn = QuanNhan.query.get(chuyen.quan_nhan_id)
    if not qn:
        flash('Không tìm thấy thông tin quân nhân.', 'danger')
        return redirect(url_for('personnel.incoming_transfers'))

    qn.don_vi_id = chuyen.don_vi_dich_id
    db.session.add(qn)

    chuyen.trang_thai = TrangThaiChuyen.CONFIRMED
    chuyen.nguoi_xac_nhan_id = current_user.id
    chuyen.ngay_xu_ly = datetime.utcnow()
    chuyen.ghi_chu = ghi_chu

    db.session.commit()
    flash(f'Đã xác nhận tiếp nhận {qn.ho_ten} vào đơn vị.', 'success')
    return redirect(url_for('personnel.incoming_transfers'))


@personnel_bp.route('/transfers/<int:transfer_id>/reject', methods=['POST'])
@login_required
@unit_user_required
def reject_transfer(transfer_id):
    """Target unit rejects the transfer request."""
    chuyen = ChuyenDonVi.query.get_or_404(transfer_id)

    if chuyen.don_vi_dich_id != current_user.don_vi_id:
        flash('Bạn không có quyền xử lý yêu cầu này.', 'danger')
        return redirect(url_for('personnel.incoming_transfers'))

    if chuyen.trang_thai != TrangThaiChuyen.PENDING:
        flash('Yêu cầu này đã được xử lý rồi.', 'warning')
        return redirect(url_for('personnel.incoming_transfers'))

    ghi_chu = request.form.get('ghi_chu', '').strip()

    chuyen.trang_thai = TrangThaiChuyen.REJECTED
    chuyen.nguoi_xac_nhan_id = current_user.id
    chuyen.ngay_xu_ly = datetime.utcnow()
    chuyen.ghi_chu = ghi_chu

    db.session.commit()
    flash(f'Đã từ chối yêu cầu chuyển {chuyen.quan_nhan.ho_ten}.', 'info')
    return redirect(url_for('personnel.incoming_transfers'))


@personnel_bp.route('/transfers/<int:transfer_id>/cancel', methods=['POST'])
@login_required
@unit_user_required
def cancel_transfer(transfer_id):
    """Source unit cancels their own pending transfer request."""
    chuyen = ChuyenDonVi.query.get_or_404(transfer_id)

    if chuyen.don_vi_nguon_id != current_user.don_vi_id:
        flash('Bạn không có quyền hủy yêu cầu này.', 'danger')
        return redirect(url_for('personnel.incoming_transfers'))

    if chuyen.trang_thai != TrangThaiChuyen.PENDING:
        flash('Yêu cầu này đã được xử lý, không thể hủy.', 'warning')
        return redirect(url_for('personnel.incoming_transfers'))

    db.session.delete(chuyen)
    db.session.commit()
    flash('Đã hủy yêu cầu chuyển đơn vị.', 'success')
    return redirect(url_for('personnel.incoming_transfers'))


# ─────────────────────────────────────────────────────────────────────────────
#  CHUYỂN VÙNG
# ─────────────────────────────────────────────────────────────────────────────

@personnel_bp.route('/chuyen-vung')
@login_required
def list_chuyen_vung():
    """Danh sách quân nhân chuyển vùng của đơn vị."""
    from app.models.user import Role as _Role
    if current_user.role not in (_Role.UNIT_USER, _Role.ADMIN):
        from flask import abort
        abort(403)

    if current_user.role == _Role.ADMIN:
        # Admin sees all chuyen vung across all units
        personnel = QuanNhan.query.filter_by(is_chuyen_vung=True).order_by(
            QuanNhan.ngay_chuyen_vung.desc(), QuanNhan.ho_ten).all()
    else:
        if not current_user.don_vi:
            flash('Tài khoản chưa được gán đơn vị.', 'warning')
            return redirect(url_for('dashboard.index'))
        personnel = QuanNhan.query.filter_by(
            don_vi_id=current_user.don_vi_id,
            is_chuyen_vung=True,
        ).order_by(QuanNhan.ngay_chuyen_vung.desc(), QuanNhan.ho_ten).all()

    return render_template('personnel/chuyen_vung.html', personnel=personnel)


@personnel_bp.route('/<int:id>/chuyen-vung', methods=['POST'])
@login_required
@unit_user_required
def mark_chuyen_vung(id):
    """Đánh dấu quân nhân chuyển vùng."""
    qn = QuanNhan.query.get_or_404(id)
    if qn.don_vi_id != current_user.don_vi_id:
        flash('Không có quyền thao tác với quân nhân này.', 'danger')
        return redirect(url_for('personnel.list_personnel'))

    if qn.is_chuyen_vung:
        flash(f'{qn.ho_ten} đã ở danh sách chuyển vùng rồi.', 'info')
        return redirect(url_for('personnel.detail_personnel', id=id))

    qn.is_chuyen_vung = True
    qn.ngay_chuyen_vung = datetime.utcnow()
    db.session.commit()
    flash(f'Đã chuyển {qn.ho_ten} sang danh sách quân nhân chuyển vùng.', 'success')
    return redirect(url_for('personnel.list_chuyen_vung'))


@personnel_bp.route('/<int:id>/huy-chuyen-vung', methods=['POST'])
@login_required
@unit_user_required
def unmark_chuyen_vung(id):
    """Hủy đánh dấu chuyển vùng — đưa quân nhân trở lại danh sách thường."""
    qn = QuanNhan.query.get_or_404(id)
    if qn.don_vi_id != current_user.don_vi_id:
        flash('Không có quyền thao tác với quân nhân này.', 'danger')
        return redirect(url_for('personnel.list_chuyen_vung'))

    qn.is_chuyen_vung = False
    qn.ngay_chuyen_vung = None
    db.session.commit()
    flash(f'Đã đưa {qn.ho_ten} trở lại danh sách quân nhân.', 'success')
    return redirect(url_for('personnel.list_chuyen_vung'))


@personnel_bp.route('/bulk-action', methods=['POST'])
@login_required
@unit_user_required
def bulk_action():
    """Bulk action: đổi đối tượng / chuyển vùng / chuyển đơn vị cho nhiều quân nhân."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': 'Dữ liệu không hợp lệ'}), 400

    action = data.get('action')
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'Không có quân nhân nào được chọn'}), 400

    # Validate all IDs belong to current unit
    qn_list = QuanNhan.query.filter(
        QuanNhan.id.in_(ids),
        QuanNhan.don_vi_id == current_user.don_vi_id,
        QuanNhan.is_active == True,
    ).all()
    if len(qn_list) != len(ids):
        return jsonify({'success': False, 'message': 'Một số quân nhân không thuộc đơn vị của bạn'}), 403

    if action == 'doi_tuong':
        doi_tuong = data.get('doi_tuong', '').strip()
        if not doi_tuong:
            return jsonify({'success': False, 'message': 'Chưa chọn đối tượng'}), 400
        for qn in qn_list:
            qn.doi_tuong = doi_tuong
        db.session.commit()
        return jsonify({'success': True, 'message': f'Đã cập nhật đối tượng cho {len(qn_list)} quân nhân'})

    elif action == 'chuyen_vung':
        for qn in qn_list:
            qn.is_chuyen_vung = True
            qn.ngay_chuyen_vung = datetime.utcnow()
        db.session.commit()
        return jsonify({'success': True, 'message': f'Đã chuyển vùng {len(qn_list)} quân nhân'})

    elif action == 'chuyen_don_vi':
        don_vi_id = data.get('don_vi_id')
        ly_do = data.get('ly_do', '').strip()
        if not don_vi_id:
            return jsonify({'success': False, 'message': 'Chưa chọn đơn vị đích'}), 400
        from app.models.unit import DonVi as _DonVi
        target_unit = _DonVi.query.get(don_vi_id)
        if not target_unit:
            return jsonify({'success': False, 'message': 'Đơn vị đích không tồn tại'}), 404
        for qn in qn_list:
            chuyen = ChuyenDonVi(
                quan_nhan_id=qn.id,
                don_vi_nguon_id=qn.don_vi_id,
                don_vi_dich_id=don_vi_id,
                ly_do=ly_do or None,
                trang_thai=TrangThaiChuyen.PENDING,
                nguoi_tao_id=current_user.id,
            )
            db.session.add(chuyen)
        db.session.commit()
        return jsonify({'success': True, 'message': f'Đã tạo yêu cầu chuyển đơn vị cho {len(qn_list)} quân nhân'})

    elif action == 'delete':
        from datetime import datetime as _dt
        for qn in qn_list:
            qn.is_active = False
            try:
                qn.is_deleted = True
                qn.deleted_at = _dt.utcnow()
                qn.deleted_by_id = current_user.id
            except Exception:
                pass
        db.session.commit()
        return jsonify({'success': True, 'message': f'Đã xóa {len(qn_list)} quân nhân'})

    else:
        return jsonify({'success': False, 'message': 'Hành động không hợp lệ'}), 400


@personnel_bp.route('/units-json')
@login_required
@unit_user_required
def get_units_json():
    """Return list of all active units for transfer dropdown."""
    from app.models.unit import DonVi as _DonVi
    units = _DonVi.query.filter_by(is_active=True).order_by(_DonVi.ten_don_vi).all()
    return jsonify({'units': [{'id': u.id, 'ten_don_vi': u.ten_don_vi} for u in units]})


# ─── Danh sách xóa (soft-delete) ────────────────────────────────────────────

@personnel_bp.route('/deleted')
@login_required
@unit_user_required
def deleted_personnel():
    """Unit user xem danh sách đã xóa của đơn vị mình (chỉ view, không thao tác)."""
    if not current_user.don_vi:
        flash('Tài khoản chưa được gán đơn vị.', 'warning')
        return redirect(url_for('dashboard.index'))
    try:
        deleted_list = QuanNhan.query.filter(
            QuanNhan.don_vi_id == current_user.don_vi_id,
            QuanNhan.is_deleted == True
        ).order_by(QuanNhan.deleted_at.desc()).all()
    except Exception:
        deleted_list = []
    return render_template('personnel/deleted_list.html', deleted_list=deleted_list)


@personnel_bp.route('/deleted/<int:id>/restore', methods=['POST'])
@login_required
def restore_personnel(id):
    """Chỉ admin mới được khôi phục."""
    if current_user.role not in (Role.ADMIN, Role.SUPER_ADMIN):
        flash('Không có quyền thực hiện thao tác này.', 'danger')
        return redirect(url_for('personnel.deleted_personnel'))
    qn = QuanNhan.query.get_or_404(id)
    qn.is_active = True
    try:
        qn.is_deleted = False
        qn.deleted_at = None
        qn.deleted_by_id = None
    except Exception:
        pass
    db.session.commit()
    flash(f'Đã khôi phục: {qn.ho_ten}', 'success')
    return redirect(url_for('admin.admin_deleted_personnel'))


@personnel_bp.route('/deleted/<int:id>/hard-delete', methods=['POST'])
@login_required
def hard_delete_personnel(id):
    """Chỉ admin mới được xóa hẳn."""
    if current_user.role not in (Role.ADMIN, Role.SUPER_ADMIN):
        flash('Không có quyền thực hiện thao tác này.', 'danger')
        return redirect(url_for('personnel.deleted_personnel'))
    qn = QuanNhan.query.get_or_404(id)
    ho_ten = qn.ho_ten
    db.session.delete(qn)
    db.session.commit()
    flash(f'Đã xóa vĩnh viễn: {ho_ten}', 'danger')
    return redirect(url_for('admin.admin_deleted_personnel'))
